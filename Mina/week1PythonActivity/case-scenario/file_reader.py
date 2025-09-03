import os
import sys
import openpyxl

filepath = r"C:\Users\WJ258LY\Downloads\PythonTraining 1.xlsx"

class file_reader:
    def __init__(self):
        self.success = False
        self.content = None

    def read_file(self):
        self.success = True
        try:
            wb = openpyxl.load_workbook(filepath)
            print("Sheet names:", wb.sheetnames)  # Debug: print all sheet names
            # Select Sheet 2 by index (second sheet)
            ws = wb.worksheets[1]  # 0 is first sheet, 1 is second
            print("Selected sheet:", ws.title)    # Debug: print selected sheet name
            self.content = []
            for row in ws.iter_rows(values_only=True):
                self.content.append(list(row))
        except Exception as e:
            self.success = False
            print(f"An error occurred while reading the file: {e}")
        return self.success

    def get_content(self):
        return self.content

#[END] - Class Bodyimport os